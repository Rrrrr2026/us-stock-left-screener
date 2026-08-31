#!/usr/bin/env python3
"""
可选：用 data/screen.json 生成 output/筛选明细.xlsx（含公式的明细表 + 说明页）。
依赖: openpyxl   (pip install openpyxl)
上行空间三列为公式（=目标价/现价-1），在 Excel/WPS 中打开会自动计算。
"""
import json
import datetime
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

BASE = pathlib.Path(__file__).parent


def rating(mark: float) -> str:
    return ("强烈买入" if mark <= 1.5 else "买入" if mark <= 2.5
            else "持有" if mark <= 3.5 else "卖出" if mark <= 4.5 else "强烈卖出")


def main() -> None:
    data = json.load(open(BASE / "data" / "screen.json"))
    sectors = data["sectors"]
    # [sym,name,sec,close,tlow,tavg,thigh,tmed,mark,ntotal,nsb,nb,nh,ns,nss,mcapB] 按 min 上行空间降序
    rows = sorted(data["rows"], key=lambda r: -(r[4] / r[3]))
    fetched = data["fetchedAt"][:16].replace("T", " ")

    wb = Workbook()
    ws = wb.active
    ws.title = "筛选结果"
    headers = ["代码", "公司名称", "板块", "现价($)", "目标价Min", "目标价Avg", "目标价Max", "目标价中位",
               "Min上行空间", "Avg上行空间", "Max上行空间", "分析师数", "综合评级",
               "强烈买入", "买入", "持有", "卖出", "强烈卖出", "市值(十亿$)"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body = Font(name="Arial", size=10)
    for i, r in enumerate(rows, start=2):
        vals = [r[0], r[1], sectors[r[2]], r[3], r[4], r[5], r[6], r[7] if r[7] is not None else "",
                f"=E{i}/D{i}-1", f"=F{i}/D{i}-1", f"=G{i}/D{i}-1",
                r[9], rating(r[8]), r[10], r[11], r[12], r[13], r[14], r[15]]
        for j, v in enumerate(vals, 1):
            ws.cell(i, j, v).font = body

    n = len(rows) + 1
    fmts = {4: "0.00", 5: "0.00", 6: "0.00", 7: "0.00", 8: "0.00",
            9: "0.0%", 10: "0.0%", 11: "0.0%", 19: "#,##0.00"}
    for col, fmt in fmts.items():
        for i in range(2, n + 1):
            ws.cell(i, col).number_format = fmt
    for j, w in enumerate([14, 34, 20, 9, 10, 10, 10, 10, 11, 11, 11, 9, 10, 8, 7, 7, 7, 9, 11], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:S{n}"
    ws.conditional_formatting.add(
        f"I2:I{n}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                                color="63A5DB", showValue=True))

    info = wb.create_sheet("说明")
    info.column_dimensions["A"].width = 110
    for i, t in enumerate([
        "美股筛选：分析师目标价 Min / Avg / Max 全部高于现价",
        "",
        f"数据抓取：{fetched} UTC · 来源 TradingView（与个股 Forecasts 页同源）· 价格为最近一个交易日收盘",
        f"全市场有 ≥5 位分析师评级的股票 {data.get('totalCovered','?')} 只，其中 {len(rows)} 只满足条件。",
        "",
        "条件：NASDAQ/NYSE/AMEX 主要上市普通股与 ADR（美元计价，不含 OTC）；评级分析师 ≥5 位；",
        "     最低一年期目标价严格高于最新收盘价（Min 在上方 ⟹ Min/Avg/Max 全部在上方）。",
        "上行空间三列为公式 =目标价/现价-1，改动现价列会自动重算。",
        "「分析师数」为给出评级的人数，与给出目标价的人数可能略有出入。",
        "评级映射：均值 ≤1.5 强烈买入，≤2.5 买入，≤3.5 持有，≤4.5 卖出。",
        "分析师目标价整体偏乐观；低价小盘股常因个别极端目标价而排名靠前，谨慎参考。不构成投资建议。",
    ], 1):
        info.cell(i, 1, t).font = Font(name="Arial", size=10)
    info.cell(1, 1).font = Font(name="Arial", size=12, bold=True)

    out = BASE / "output" / "筛选明细.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"✓ {out}（{len(rows)} 行）")


if __name__ == "__main__":
    main()
